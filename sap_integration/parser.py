import json
import os
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


SCHEMA_VERSION = 1
TRANSACTION_HEADERS = {
    "Fonds",
    "Name des Geschäftspartners",
    "Betrag",
    "Belegkopftext",
    "Positionstext",
}
BUDGET_HEADERS = {"Fonds", "Betrag"}


def parse_downloaded_reports(data_dir, year, fund_numbers):
    data_dir = Path(data_dir)
    raw_dir = data_dir / "raw" / str(year)
    report_paths = {
        "budget": raw_dir / "budget.xlsx",
        "actual": raw_dir / "actual.xlsx",
        "commitments": raw_dir / "commitments.xlsx",
    }
    for report_name, report_path in report_paths.items():
        if not report_path.is_file():
            raise FileNotFoundError(
                f"SAP-Export für {report_name} fehlt: {report_path}"
            )

    normalized_funds = {_normalize_fund_number(value) for value in fund_numbers}
    normalized_funds.discard("")
    funds = {
        fund_number: {
            "fund_number": fund_number,
            "has_budget": False,
            "budget": Decimal("0"),
            "actual_total": Decimal("0"),
            "commitments_total": Decimal("0"),
            "transactions": [],
        }
        for fund_number in sorted(normalized_funds)
    }

    _parse_budget(report_paths["budget"], funds)
    _parse_transactions(report_paths["actual"], funds, "actual")
    _parse_transactions(report_paths["commitments"], funds, "commitment")

    for fund in funds.values():
        combined = fund["actual_total"] + fund["commitments_total"]
        fund["combined_total"] = combined
        fund["remaining"] = fund["budget"] - combined if fund["has_budget"] else None
        for key in (
            "budget",
            "actual_total",
            "commitments_total",
            "combined_total",
            "remaining",
        ):
            fund[key] = _decimal_string(fund[key]) if fund[key] is not None else None

    generated_at = datetime.now(timezone.utc).isoformat()
    payload = {
        "schema_version": SCHEMA_VERSION,
        "year": year,
        "generated_at": generated_at,
        "funds": funds,
    }
    target_dir = data_dir / "processed"
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f"{year}.json"
    temp_path = target_dir / f".{year}.json.tmp"
    with temp_path.open("w", encoding="utf-8") as output_file:
        json.dump(payload, output_file, ensure_ascii=False, indent=2)
        output_file.write("\n")
    os.replace(temp_path, target_path)
    return target_path


def _parse_budget(path, funds):
    for row in _iter_rows(path, BUDGET_HEADERS):
        fund = funds.get(_normalize_fund_number(row["Fonds"]))
        if fund is None:
            continue
        fund["has_budget"] = True
        fund["budget"] += _decimal(row["Betrag"])


def _parse_transactions(path, funds, transaction_type):
    for row in _iter_rows(path, TRANSACTION_HEADERS):
        fund = funds.get(_normalize_fund_number(row["Fonds"]))
        if fund is None:
            continue

        amount = _decimal(row["Betrag"])
        position = " ".join(
            part
            for part in (
                _text(row.get("Belegkopftext")),
                _text(row.get("Positionstext")),
            )
            if part
        )
        fund["transactions"].append(
            {
                "type": transaction_type,
                "business_partner": _text(row.get("Name des Geschäftspartners")),
                "position": position,
                "amount": _decimal_string(amount),
                "booking_date": _date_string(row.get("Buchungsdatum")),
            }
        )
        total_key = "actual_total" if transaction_type == "actual" else "commitments_total"
        fund[total_key] += amount


def _iter_rows(path, required_headers):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as error:
            raise ValueError(f"Leerer SAP-Export: {Path(path).name}") from error
        headers = [_text(value) for value in raw_headers]
        missing = required_headers - set(headers)
        if missing:
            raise ValueError(
                f"Fehlende Spalten in {Path(path).name}: {', '.join(sorted(missing))}"
            )
        for values in rows:
            yield dict(zip(headers, values))
    finally:
        workbook.close()


def _normalize_fund_number(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    normalized = str(value).strip().replace("€", "").replace(" ", "")
    if "," in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation as error:
        raise ValueError(f"Ungültiger SAP-Betrag: {value!r}") from error


def _decimal_string(value):
    return format(value.quantize(Decimal("0.01")), "f")


def _text(value):
    return "" if value is None else str(value).strip()


def _date_string(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _text(value) or None
